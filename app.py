import os
import io
import datetime
from pathlib import Path

import pandas as pd
from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    session,
    send_file,
    flash,
)
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from werkzeug.security import generate_password_hash, check_password_hash

# -----------------------------------------------------------------------------
# Flask app setup
# -----------------------------------------------------------------------------
app = Flask(__name__)
app.secret_key = "change_this_to_a_strong_random_secret"  # TODO: change in prod

# File storage & retention
BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

RETENTION_DAYS = 7  # keep last N days of uploads

# Single user credentials
USERNAME = "admin"
PASSWORD_HASH = generate_password_hash("changeme123")  # TODO: change password

# Global in-memory report
LATEST_REPORT_DF: pd.DataFrame | None = None

# Agent → email mapping
AGENT_EMAILS = {
    "Jason Canales": "jason.canales@way.com",
    "Joseph Parker": "joseph.parker@way.com",
    "Ayrton Oneal": "ayrton.oneal@way.com",
    "Mike Zimmerman": "mike.zimmerman@way.com",
    "Felecia Boswell": "felecia.boswell@wayinsured.com",
    "Drew Backus": "drew.backus@way.com",
    "Joe Hodges": "joseph.hodges@way.com",
    "Matthew Laushman": "matthew.laushman@way.com",
    "Larry Johnson": "larry.johnson@way.com",
    "Matthew Mandra": "matthew.mandra@way.com",
    "Lee Oday": "lee.oday@way.com",
    "Bryant Hamman": "bryant.hamman@way.com",
    "Kenyatta Rutland": "kenyatta.rutland@way.com",
    "Carlos Villanueva": "carlos.villanueva@way.com",
    "Amr Mohamed": "amr.m@way.com",
    "Lucas Vargas": "lucas.vargas@way.com",
    "Mike Alvarez": "mike.alvarez@way.com",
    "Francesca Cristantiello": "francesca.cristantiello@way.com",
    "Alejandro Gonzalez": "alejandro.gonzalez@way.com",
    "Angie Covilla": "angie.covilla@way.com",
    "Laura Gutierrez": "laura.gutierrez@way.com",
    "Yulieth Jimenez": "yulieth.jimenez@way.com",
    "Julian Hernandez": "julian.hernandez@way.com",
    "Laura Vasquez": "laura.vasquez@way.com",
    "Linda Mayorquin": "linda.mayorquin@way.com",
    "Key Kim": "key.kim@way.com",
}


# -----------------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------------
def prune_old_uploads() -> None:
    """Delete uploaded CSV files older than RETENTION_DAYS."""
    cutoff = datetime.datetime.now() - datetime.timedelta(days=RETENTION_DAYS)
    for path in UPLOAD_DIR.glob("*.csv"):
        mtime = datetime.datetime.fromtimestamp(path.stat().st_mtime)
        if mtime < cutoff:
            try:
                path.unlink()
            except Exception:
                # best-effort delete, ignore errors
                pass


def login_required(f):
    """Decorator to require login on a route."""
    from functools import wraps

    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)

    return wrapper


# -----------------------------------------------------------------------------
# Auth routes
# -----------------------------------------------------------------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "")
        password = request.form.get("password", "")

        if username == USERNAME and check_password_hash(PASSWORD_HASH, password):
            session["logged_in"] = True
            return redirect(url_for("upload_files"))
        else:
            flash("Invalid username or password", "danger")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# -----------------------------------------------------------------------------
# Upload + generate report
# -----------------------------------------------------------------------------
@app.route("/", methods=["GET", "POST"])
@login_required
def upload_files():
    if request.method == "POST":
        new_tile_file = request.files.get("new_tile")
        call_log_file = request.files.get("call_log")

        if not new_tile_file or not call_log_file:
            flash("Please upload both new_tile.csv and call_log.csv", "warning")
            return redirect(url_for("upload_files"))

        # timestamp for filenames
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        new_tile_path = UPLOAD_DIR / f"{ts}_new_tile.csv"
        call_log_path = UPLOAD_DIR / f"{ts}_call_log.csv"

        try:
            # save uploads
            new_tile_file.save(new_tile_path)
            call_log_file.save(call_log_path)

            # prune old
            prune_old_uploads()

            # load into pandas
            new_tile_df = pd.read_csv(new_tile_path)
            call_log_df = pd.read_csv(call_log_path)
        except Exception as e:
            flash(f"Error processing CSV files: {e}", "danger")
            return redirect(url_for("upload_files"))

        global LATEST_REPORT_DF
        LATEST_REPORT_DF = generate_report(new_tile_df, call_log_df)

        flash("Report generated successfully!", "success")
        return redirect(url_for("dashboard"))

    return render_template("upload.html")


# -----------------------------------------------------------------------------
# Dashboard + agent detail
# -----------------------------------------------------------------------------
@app.route("/dashboard")
@login_required
def dashboard():
    global LATEST_REPORT_DF
    if LATEST_REPORT_DF is None:
        flash("No report generated yet. Please upload files first.", "info")
        return redirect(url_for("upload_files"))

    rows = LATEST_REPORT_DF.to_dict(orient="records")

    agents = LATEST_REPORT_DF["Agent Name"].tolist()
    total_outbound = LATEST_REPORT_DF["Total Outbound Calls"].tolist()
    manual_outbound = LATEST_REPORT_DF["No. of Manual Outbound Calls"].tolist()
    inbound_calls = LATEST_REPORT_DF["No. of Inbound Calls"].tolist()

    return render_template(
        "dashboard.html",
        rows=rows,
        agents=agents,
        total_outbound=total_outbound,
        manual_outbound=manual_outbound,
        inbound_calls=inbound_calls,
    )


@app.route("/agent/<name>")
@login_required
def agent_detail(name):
    global LATEST_REPORT_DF
    if LATEST_REPORT_DF is None:
        flash("No report generated yet.", "info")
        return redirect(url_for("upload_files"))

    # agent name from URL (Flask already decodes %20 -> space)
    agent_name = name
    agent_row = LATEST_REPORT_DF[LATEST_REPORT_DF["Agent Name"] == agent_name]

    if agent_row.empty:
        flash(f"No data found for agent: {agent_name}", "warning")
        return redirect(url_for("dashboard"))

    row = agent_row.iloc[0].to_dict()
    return render_template("agent.html", row=row)


# -----------------------------------------------------------------------------
# Download Excel with formatting
# -----------------------------------------------------------------------------
@app.route("/download")
@login_required
def download_report():
    global LATEST_REPORT_DF
    if LATEST_REPORT_DF is None:
        flash("No report generated yet.", "info")
        return redirect(url_for("upload_files"))

    # create Excel in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        LATEST_REPORT_DF.to_excel(writer, index=False, sheet_name="Report")

    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active

    # color Total Calls (column E)
    total_calls_col = 5  # E
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=total_calls_col)
        value = cell.value or 0
        if value < 75:
            cell.fill = red_fill
        else:
            cell.fill = green_fill

    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    return send_file(
        final_output,
        as_attachment=True,
        download_name="agent_summary_fixed.xlsx",
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )


# -----------------------------------------------------------------------------
# Core report logic (your agreed business rules)
# -----------------------------------------------------------------------------
def generate_report(new_tile: pd.DataFrame, call_log: pd.DataFrame) -> pd.DataFrame:
    """
    - Total Outbound Calls        = Outbound Calls Completed (tile)
    - No. of Manual Outbound     = count of agent email in call_log
    - No. of Inbound Calls       = Inbound Calls Completed (tile)
    - Total Talk Time (min)      = Total Talk Time (min) (tile)
    - If agent not in tile       = tile metrics 0, but manual outbound still from log
    - Total Calls                = Total Outbound Calls + No. of Inbound Calls
    """

    # Normalize agent names in tile
    new_tile = new_tile.copy()
    new_tile["Agent Name"] = (
        new_tile["Agent Name"].astype(str).str.strip().str.title()
    )

    # Filter tile to only agents in mapping
    tile_filtered = new_tile[new_tile["Agent Name"].isin(AGENT_EMAILS.keys())].copy()

    # If multiple rows per agent (different teams), keep one (they are same for that date)
    tile_filtered = tile_filtered.drop_duplicates(subset=["Agent Name"], keep="last")

    # Keep needed tile columns
    tile_filtered = tile_filtered[
        [
            "Agent Name",
            "Outbound Calls Completed",
            "Inbound Calls Completed",
            "Total Talk Time (min)",
        ]
    ]

    # Manual outbound from call_log
    call_log = call_log.copy()
    call_log["Handling Agent Email"] = (
        call_log["Handling Agent Email"].astype(str).str.lower().str.strip()
    )
    email_counts = call_log["Handling Agent Email"].value_counts().to_dict()

    # Base DF with all agents from mapping
    final_df = pd.DataFrame({"Agent Name": list(AGENT_EMAILS.keys())})
    final_df["Agent Email"] = final_df["Agent Name"].map(AGENT_EMAILS).str.lower()

    # Merge tile metrics
    final_df = final_df.merge(tile_filtered, on="Agent Name", how="left")

    # Fill NaNs for agents missing in tile
    final_df["Outbound Calls Completed"] = final_df["Outbound Calls Completed"].fillna(0)
    final_df["Inbound Calls Completed"] = final_df["Inbound Calls Completed"].fillna(0)
    final_df["Total Talk Time (min)"] = final_df["Total Talk Time (min)"].fillna(0)

    # 1) Total Outbound Calls = tile outbound
    final_df["Total Outbound Calls"] = final_df["Outbound Calls Completed"]

    # 2) No. of Manual Outbound Calls = count from call_log
    final_df["No. of Manual Outbound Calls"] = final_df["Agent Email"].map(
        lambda e: email_counts.get(e, 0)
    )

    # 3) No. of Inbound Calls = tile inbound
    final_df["No. of Inbound Calls"] = final_df["Inbound Calls Completed"]

    # 4) Talk time
    final_df["Total Talk Time (min)"] = final_df["Total Talk Time (min)"].round(2)

    # 5) Total Calls
    final_df["Total Calls"] = (
        final_df["Total Outbound Calls"] + final_df["No. of Inbound Calls"]
    )

    # Drop helper columns from tile
    final_df = final_df.drop(
        columns=["Agent Email", "Outbound Calls Completed", "Inbound Calls Completed"]
    )

    # Column order
    final_df = final_df[
        [
            "Agent Name",
            "Total Outbound Calls",
            "No. of Manual Outbound Calls",
            "No. of Inbound Calls",
            "Total Calls",
            "Total Talk Time (min)",
        ]
    ]

    return final_df


# -----------------------------------------------------------------------------
# Run app
# -----------------------------------------------------------------------------
if __name__ == "__main__":
    # Dev mode only; use gunicorn/uwsgi + nginx in production
    app.run(debug=True)
